from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ParsedAspect:
    name: str
    target: str
    severity: str
    phase: str
    prompt_template: str


EXPECTED_HEADERS = ("観点名", "対象", "重要度", "工程", "レビュー指示文")
VALID_TARGETS = {"UI", "SS", "UI×SS"}
SEVERITY_MAP = {
    "高": "high",
    "中": "mid",
    "低": "low",
    "high": "high",
    "mid": "mid",
    "medium": "mid",
    "low": "low",
}


def parse_aspect_file(path: Path) -> list[ParsedAspect]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_excel_aspects(path)
    if suffix == ".txt":
        return parse_txt_aspects(path)
    raise ValueError(f"Unsupported aspect file extension: {suffix}")


def parse_excel_aspects(path: Path) -> list[ParsedAspect]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "観点" not in workbook.sheetnames:
            raise ValueError('Excel aspect file must contain sheet "観点"')
        sheet = workbook["観点"]
        _validate_excel_headers(sheet)
        aspects: list[ParsedAspect] = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=5, values_only=True), start=2):
            values = [_clean_cell(value) for value in row]
            if not any(values):
                continue
            aspects.append(_build_aspect(values, row_label=f"row {row_index}", cell_prefix=f"row {row_index} column "))
        if not aspects:
            raise ValueError("Excel aspect file has no aspect rows")
        return aspects
    finally:
        workbook.close()


def parse_txt_aspects(path: Path) -> list[ParsedAspect]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    aspects: list[ParsedAspect] = []
    for line_index, line in enumerate(lines, start=1):
        stripped = line.strip()
        if not stripped:
            continue
        parts = [part.strip() for part in stripped.split("|")]
        if line_index == 1 and tuple(parts) == EXPECTED_HEADERS:
            continue
        if len(parts) != 5:
            raise ValueError(f"line {line_index}: expected 5 pipe-delimited fields, got {len(parts)}")
        aspects.append(_build_aspect(parts, row_label=f"line {line_index}", cell_prefix=f"line {line_index} field "))
    if not aspects:
        raise ValueError("TXT aspect file has no aspect rows")
    return aspects


def _validate_excel_headers(sheet: Worksheet) -> None:
    values = [_clean_cell(sheet.cell(row=1, column=index).value) for index in range(1, 6)]
    for index, (actual, expected) in enumerate(zip(values, EXPECTED_HEADERS, strict=True), start=1):
        if actual != expected:
            column = chr(ord("A") + index - 1)
            raise ValueError(
                f"row 1 column {column}: expected header {expected!r}, got {actual!r}"
            )


def _build_aspect(values: list[str], row_label: str, cell_prefix: str) -> ParsedAspect:
    for index, (value, header) in enumerate(zip(values, EXPECTED_HEADERS, strict=True), start=1):
        if not value:
            label = chr(ord("A") + index - 1) if cell_prefix.endswith("column ") else str(index)
            raise ValueError(f"{cell_prefix}{label} ({header}) is required")

    target = _normalize_target(values[1], row_label)
    severity = _normalize_severity(values[2], row_label)
    return ParsedAspect(
        name=values[0],
        target=target,
        severity=severity,
        phase=values[3],
        prompt_template=values[4],
    )


def _normalize_target(value: str, row_label: str) -> str:
    normalized = value.replace(" ", "").replace("x", "×").replace("X", "×")
    if normalized not in VALID_TARGETS:
        raise ValueError(f"{row_label}: target must be one of UI, SS, UI×SS, got {value!r}")
    return normalized


def _normalize_severity(value: str, row_label: str) -> str:
    normalized = SEVERITY_MAP.get(value.lower() if value.isascii() else value)
    if normalized is None:
        raise ValueError(f"{row_label}: severity must be one of 高, 中, 低, got {value!r}")
    return normalized


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
