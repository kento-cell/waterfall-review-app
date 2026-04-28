from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.parsers.base import ParsedArtifact


class XlsxParser:
    def parse(self, path: Path) -> ParsedArtifact:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        try:
            for sheet in workbook.worksheets:
                lines.append(f"# Sheet: {sheet.title}")
                for row_index, row in enumerate(sheet.iter_rows(), start=1):
                    for column_index, cell in enumerate(row, start=1):
                        value = "" if cell.value is None else str(cell.value)
                        coordinate = f"{get_column_letter(column_index)}{row_index}"
                        lines.append(f"{sheet.title}!{coordinate}: {value}")
        finally:
            workbook.close()
        return ParsedArtifact(content="\n".join(lines))
