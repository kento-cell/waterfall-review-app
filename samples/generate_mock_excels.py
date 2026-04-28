"""
AIレビュー支援ツール 検証用モックデータ生成 v3 (本格 SI 品質)
受領フォーマット精査結果反映:
  - UI: UI_0029 (機能概要書) 形式 — 番号階層 (1)→①→(a)→(ⅰ) を採用、画面詳細レイアウト
  - SS: SS_バッチ処理仕様 形式 — 番号付き処理ブロック (取得/更新/削除条件 詳細), メッセージID内蔵
出力: スクリプトと同じディレクトリ (samples/)
ライブラリ: openpyxl (新規ファイル作成のみ)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent

# ====== スタイル定数 ======
HEADER_FILL = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="3D6098")
SUBHEAD_FONT = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
META_FILL = PatternFill("solid", fgColor="D9E2F3")
META_FONT = Font(name="游ゴシック", size=9, bold=True)
SECTION_FONT = Font(name="游明朝", size=12, bold=True, color="1B3A6B")
SUB_FONT = Font(name="游明朝", size=10, bold=True, color="0B1F3F")
BODY_FONT = Font(name="游ゴシック", size=9)
BODY_BOLD = Font(name="游ゴシック", size=9, bold=True)
MONO_FONT = Font(name="Consolas", size=9)
WARN_FONT = Font(name="游ゴシック", size=9, italic=True, color="C62828")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_MERGE = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ====== 定型ヘッダ (UI_0029 表紙風) ======
def add_doc_header(ws, sheet_label, project, phase, subsystem, author, created, updater, updated):
    rows = [
        ("プロジェクト名", "局面", "サブシステム名", sheet_label, "作成者", "作成日", "更新者", "更新日"),
        (project, phase, subsystem, "", author, created, updater, updated),
    ]
    cols = ["A", "C", "E", "H", "L", "N", "P", "R"]
    widths = [15, 10, 16, 22, 8, 12, 8, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c_letter, label in zip(cols, rows[0]):
        cell = ws[f"{c_letter}1"]
        cell.value = label
        cell.font = META_FONT
        cell.fill = META_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for c_letter, val in zip(cols, rows[1]):
        cell = ws[f"{c_letter}2"]
        cell.value = val
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20


# ============================================================================
# UI_機能概要書 (本格版)
# ============================================================================
def gen_ui():
    wb = Workbook()
    PROJ = "受注管理システム改修"
    META = (PROJ, "基本設計", "受注サブシステム", "山田", "2026-04-15", "鈴木", "2026-04-25")

    # === 表紙 ===
    ws = wb.active
    ws.title = "表紙"
    add_doc_header(ws, "表紙", *META)
    ws["B5"] = "受注管理システム改修 機能概要書"
    ws["B5"].font = Font(name="游明朝", size=20, bold=True, color="1B3A6B")
    ws.merge_cells("B5:R6")
    ws["B5"].alignment = CENTER
    ws["B9"] = "概要"
    ws["B9"].font = SECTION_FONT
    ws["B10"] = ("ABC社向け受注管理システムに以下の機能を追加・改修する。\n"
                 "  ・受注登録、検索、承認、削除、メール通知、出力 (合計 7 機能)\n"
                 "  ・10万円超の受注は承認フラグを自動セット (運用効率化)\n"
                 "  ・既存顧客マスタ参照を維持しつつ、受注伝票の項目を拡充")
    ws["B10"].font = BODY_FONT
    ws.merge_cells("B10:R14")
    ws["B10"].alignment = LEFT_TOP
    ws["B16"] = "対象顧客"
    ws["B16"].font = SUB_FONT
    ws["B17"] = "ABC株式会社"
    ws["B17"].font = BODY_FONT
    ws["B18"] = "対象システム"
    ws["B18"].font = SUB_FONT
    ws["B19"] = "受注サブシステム (販売管理基盤の一部)"
    ws["B19"].font = BODY_FONT

    # === 改訂履歴 ===
    ws = wb.create_sheet("改訂履歴")
    add_doc_header(ws, "改訂履歴", *META)
    ws["B4"] = "改訂履歴"
    ws["B4"].font = SECTION_FONT
    headers = ["No", "版", "日付", "更新者", "改訂内容", "備考"]
    cw = [4, 6, 12, 8, 50, 18]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    rev = [
        (1, "0.1", "2026-04-15", "山田", "初版作成 (機能・帳票一覧、改造基本方針)", ""),
        (2, "0.5", "2026-04-18", "山田", "テーブル設計章追加 (受注/受注明細/承認履歴)", ""),
        (3, "1.0", "2026-04-20", "山田", "顧客レビュー反映 (承認閾値10万円、承認フロー追加)", "顧客承認"),
        (4, "1.1", "2026-04-22", "山田", "F-006 削除機能の文言修正 (★『適切に処理』のまま残置)", ""),
        (5, "1.2", "2026-04-25", "鈴木", "F-007 メール通知 追加、画面項目補足", ""),
    ]
    for r, row in enumerate(rev, start=7):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c <= 4 else LEFT
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === 1. 機能・帳票一覧 ===
    ws = wb.create_sheet("1.機能・帳票一覧")
    add_doc_header(ws, "機能・帳票一覧", *META)
    ws["B4"] = "1．機能・帳票一覧"
    ws["B4"].font = SECTION_FONT
    ws["C5"] = "本案件で新規追加・改修する画面、帳票、バッチ処理を以下に示す。"
    ws["C5"].font = BODY_FONT
    headers = ["No", "機能名", "プロセスID", "工場/部門", "区分", "種別", "概要", "備考"]
    cw = [4, 16, 12, 10, 8, 6, 38, 22]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    rows = [
        (1, "受注登録",     "ORDC100", "営業部",  "新規", "画面",   "顧客から受注情報を登録、自動採番", ""),
        (2, "受注検索",     "ORDC110", "営業部",  "新規", "画面",   "受注番号/顧客名/日付範囲で検索", ""),
        (3, "受注承認",     "ORDC120", "営業部長", "新規", "画面",   "金額10万円超は承認待ち、承認/差戻し", ""),
        (4, "金額計算",     "ORDS200", "共通",    "改修", "処理",   "数量×単価、税込換算 (8% / 10%)", "★税率の根拠不明"),
        (5, "受注一覧出力", "ORDR300", "営業部",  "新規", "帳票",   "Excel 形式で受注一覧出力", ""),
        (6, "受注削除",     "ORDC130", "営業部長", "新規", "画面",   "受注を削除 (適切に処理する)", "★『適切に』が曖昧"),
        (7, "メール通知",   "ORDB400", "共通",    "新規", "バッチ", "登録/承認時に営業担当へ通知", ""),
    ]
    for r, row in enumerate(rows, start=8):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 3, 4, 5, 6) else LEFT
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === 2. 改造基本方針 ===
    ws = wb.create_sheet("2.改造基本方針")
    add_doc_header(ws, "改造基本方針", *META)
    ws["B4"] = "2．改造基本方針"
    ws["B4"].font = SECTION_FONT

    method_blocks = [
        ("(1) 受注登録 (ORDC100)", [
            "・顧客マスタ参照で顧客名選択 (検索ダイアログ)。",
            "・数量×単価で金額自動計算 (税込表示)。",
            "・受注番号は自動採番 (ORD + 8桁日付 + 連番3桁)。",
            "・登録ボタン押下時、入力検証 (必須/桁数/数値範囲) を実施。",
        ]),
        ("(2) 受注承認 (ORDC120)", [
            "・受注一覧から金額10万円超の受注を抽出して表示。",
            "・「承認」ボタンで承認フラグ=true、TBCTHB05承認履歴に登録。",
            "・「差戻し」ボタンで状態を「差戻し」に更新、理由を必須入力。",
            "・承認/差戻しともに営業担当へメール通知 (バッチ ORDB400 連携)。",
        ]),
        ("(3) 金額計算 (ORDS200)", [
            "・税率は商品マスタの税区分に従う (8% 軽減税率 / 10% 標準)。",
            "・端数処理は四捨五入 (小数第1位)。",
            "・金額が10万円を超える場合は承認フラグを true にセットする。",
            "★税率の閾値は顧客と要再確認 (8%対象商品の定義が不明確)。",
        ]),
        ("(4) 受注削除 (ORDC130)", [
            "・削除ボタン押下で確認ダイアログ表示 (MSG_005)。",
            "・「OK」で論理削除 (DELETE_FLG=1) + 履歴テーブル登録。",
            "・削除後は適切に処理する。",
            "★『適切に処理する』の具体的内容を明記すること (関連レコード/在庫戻し有無)。",
        ]),
        ("(5) メール通知 (ORDB400)", [
            "・受注登録時、および承認/差戻し時に営業担当へ自動送信。",
            "・宛先は顧客マスタの担当者メールアドレスを使用。",
            "・本文テンプレートは別途仕様 (MSG_TPL_001〜003) に従う。",
        ]),
    ]
    r = 6
    for title, items in method_blocks:
        ws.cell(row=r, column=3, value=title).font = SUB_FONT
        r += 1
        for it in items:
            cell = ws.cell(row=r, column=4, value=it)
            if it.startswith("★"):
                cell.font = WARN_FONT
            else:
                cell.font = BODY_FONT
            cell.alignment = LEFT_TOP
            r += 1
        r += 1
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 70

    # === 3. テーブル設計 ===
    ws = wb.create_sheet("3.テーブル設計")
    add_doc_header(ws, "テーブル設計", *META)
    ws["B4"] = "3．テーブル設計"
    ws["B4"].font = SECTION_FONT
    ws["C5"] = "※カラムは代表的なもののみ記載。共通カラム (登録日時等) は省略。"
    ws["C5"].font = BODY_FONT

    tables = [
        ("TBCTHB03 受注ヘッダ", [
            ("PK", "ORDER_NO",     "CHAR(13)",  "受注番号 (ORD+YYYYMMDD+SEQ3)"),
            ("",   "CUSTOMER_CD",  "CHAR(8)",   "顧客コード (FK 顧客マスタ)"),
            ("",   "ORDER_DATE",   "DATE",      "受注日 (NOT NULL)"),
            ("",   "TOTAL_AMOUNT", "NUMBER(12)","合計金額 (税込)"),
            ("",   "APPROVE_FLG",  "CHAR(1)",   "承認フラグ (0:未/1:済/2:差戻)"),
            ("",   "STATUS",       "CHAR(2)",   "状態 (10:登録/20:承認待/30:承認済/90:削除)"),
            ("",   "DELETE_FLG",   "CHAR(1)",   "論理削除 (0:有効/1:削除)"),
        ]),
        ("TBCTHB04 受注明細", [
            ("PK", "ORDER_NO",     "CHAR(13)",  "受注番号 (FK 受注ヘッダ)"),
            ("PK", "LINE_NO",      "NUMBER(3)", "明細番号"),
            ("",   "ITEM_CD",      "CHAR(10)",  "商品コード"),
            ("",   "QUANTITY",     "NUMBER(5)", "数量 (1以上)"),
            ("",   "UNIT_PRICE",   "NUMBER(10)","単価 (★必須/任意指定なし)"),
            ("",   "TAX_RATE",     "NUMBER(2)", "税率 (8/10)"),
            ("",   "AMOUNT",       "NUMBER(12)","小計金額 (税込)"),
        ]),
        ("TBCTHB05 承認履歴 (新規)", [
            ("PK", "HISTORY_NO",   "NUMBER(10)","履歴番号 (シーケンス採番)"),
            ("",   "ORDER_NO",     "CHAR(13)",  "受注番号 (FK)"),
            ("",   "ACTION",       "CHAR(2)",   "10:承認/20:差戻し"),
            ("",   "ACTION_USER",  "CHAR(8)",   "実施ユーザコード"),
            ("",   "ACTION_TIME",  "TIMESTAMP", "実施日時"),
            ("",   "REASON",       "VARCHAR2(200)", "差戻し理由 (差戻し時必須)"),
        ]),
    ]
    r = 7
    for tname, cols_def in tables:
        ws.cell(row=r, column=2, value=tname).font = SUB_FONT
        r += 1
        for i, h in enumerate(["PK", "カラム名", "型/桁", "説明"], 1):
            c = ws.cell(row=r, column=2 + i, value=h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
        r += 1
        for col_def in cols_def:
            for i, v in enumerate(col_def, 1):
                cell = ws.cell(row=r, column=2 + i, value=v)
                cell.font = MONO_FONT if i in (1, 2, 3) else BODY_FONT
                cell.border = BORDER
                cell.alignment = CENTER if i in (1, 3) else LEFT
            r += 1
        r += 1
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 50

    # === 4. 画面個別詳細 (受注登録 ORDC100) ===
    ws = wb.create_sheet("4.画面詳細_受注登録")
    add_doc_header(ws, "画面詳細", *META)
    ws["B4"] = "4．画面個別詳細  (1) 受注登録画面 ORDC100"
    ws["B4"].font = SECTION_FONT

    ws["C6"] = "(1) 画面レイアウト"
    ws["C6"].font = SUB_FONT
    ws["D7"] = "①ヘッダ部"
    ws["D7"].font = BODY_BOLD
    ws["E8"] = "(a) 受注番号 (自動採番、ReadOnly)、受注日 (デフォルト=システム日付)"
    ws["E9"] = "(b) 顧客コード+顧客名 (顧客マスタ検索ダイアログから設定)"
    ws["D10"] = "②明細部"
    ws["D10"].font = BODY_BOLD
    ws["E11"] = "(a) 明細行は最大20行。行追加/削除ボタンを配置"
    ws["E12"] = "(b) 商品コードを選択すると商品マスタから商品名・単価・税率を自動セット"
    ws["E13"] = "(c) 数量入力で小計金額を自動計算 (数量×単価、税込)"
    ws["D14"] = "③合計部"
    ws["D14"].font = BODY_BOLD
    ws["E15"] = "(a) 合計金額 (各明細の小計合計、税込)"
    ws["E16"] = "(b) 承認状態表示 (合計>10万円なら『承認待ち』ラベル)"

    ws["C18"] = "(2) 画面項目定義"
    ws["C18"].font = SUB_FONT
    headers = ["項目No", "項目名", "型", "桁数", "必須", "初期値", "業務ルール"]
    cw_main = [8, 14, 8, 6, 8, 12, 36]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=20, column=2 + i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    items = [
        (1, "受注番号",     "文字",  13, "自動",     "自動採番",    "ORD+YYYYMMDD+SEQ3"),
        (2, "受注日",       "日付",  10, "必須",     "システム日付", "未来日不可"),
        (3, "顧客コード",   "文字",   8, "必須",     "",           "顧客マスタに存在すること"),
        (4, "顧客名",       "文字",  50, "自動",     "",           "顧客マスタから自動セット"),
        (5, "商品コード",   "文字",  10, "必須",     "",           "商品マスタに存在すること"),
        (6, "数量",         "数値",   5, "必須",     "1",          "1以上"),
        (7, "単価",         "数値",  10, "",         "",           "★必須/任意の指定なし"),
        (8, "税率",         "数値",   2, "自動",     "10",         "商品マスタの税区分から (8 or 10)"),
        (9, "小計金額",     "数値",  12, "自動計算", "0",          "数量×単価×(1+税率/100)"),
        (10,"合計金額",     "数値",  12, "自動",     "0",          "明細の小計合計 (税込)"),
        (11,"備考",         "文字", 200, "任意",     "",           ""),
    ]
    for r, row in enumerate(items, start=21):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=2 + c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 3, 4, 5, 6) else LEFT
            if isinstance(v, str) and "★" in v:
                cell.font = WARN_FONT
    for i, w in enumerate(cw_main, 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = w
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8

    ws["C34"] = "(3) 画面項目制御"
    ws["C34"].font = SUB_FONT
    ws["D35"] = "①初期表示"
    ws["D35"].font = BODY_BOLD
    ws["E36"] = "(a) 「メインメニュー」「キャンセル」「登録」ボタンを活性"
    ws["E37"] = "(b) 受注番号/顧客名は ReadOnly、その他項目は入力可"
    ws["D38"] = "②登録ボタン押下時"
    ws["D38"].font = BODY_BOLD
    ws["E39"] = "(a) 入力検証 (必須/桁数/数値範囲) を順次実施"
    ws["E40"] = "(b) 検証 OK の場合、TBCTHB03 受注ヘッダ + TBCTHB04 受注明細 を INSERT"
    ws["E41"] = "(c) 合計金額 > 10万円の場合、TBCTHB03.APPROVE_FLG=0、STATUS='20' (承認待ち)"
    ws["E42"] = "(d) 承認待ちの場合、メール通知バッチ用キューに enqueue (ORDB400 連携)"

    # === 5. メッセージ定義 ===
    ws = wb.create_sheet("5.メッセージ定義")
    add_doc_header(ws, "メッセージ定義", *META)
    ws["B4"] = "5．メッセージ定義"
    ws["B4"].font = SECTION_FONT
    headers = ["メッセージID", "種別", "メッセージ内容", "発生条件", "対象機能"]
    cw = [14, 8, 36, 24, 16]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    msgs = [
        ("MSG_001", "情報",   "登録しました",                          "登録成功",            "ORDC100"),
        ("MSG_002", "情報",   "削除しました",                          "削除成功",            "ORDC130"),
        ("MSG_003", "エラー", "必須項目を入力してください",            "必須未入力",          "全画面"),
        ("MSG_004", "エラー", "数量は1以上で入力してください",          "数量0以下",           "ORDC100"),
        ("MSG_005", "確認",   "削除しますか?",                         "削除ボタン押下",      "ORDC130"),
        ("MSG_006", "情報",   "承認待ちとなりました",                  "金額>10万円",         "ORDC100"),
        ("MSG_007", "情報",   "承認しました",                          "承認実行",            "ORDC120"),
        ("MSG_008", "確認",   "差戻し理由を入力してください",          "差戻し時必須",        "ORDC120"),
        ("MSG_009", "エラー", "顧客マスタに存在しない顧客コードです",   "顧客存在チェック NG", "ORDC100"),
    ]
    for r, row in enumerate(msgs, start=7):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 5) else LEFT
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = OUT / "sample_UI_機能概要書.xlsx"
    wb.save(out)
    return out


# ============================================================================
# SS_構造設計書 (本格版)
# ============================================================================
def gen_ss():
    wb = Workbook()
    META_SS = (
        "受注管理システム改修", "詳細設計", "受注サブシステム",
        "佐藤", "2026-04-22", "佐藤", "2026-04-26",
    )

    # === 変更履歴 ===
    ws = wb.active
    ws.title = "変更履歴"
    add_doc_header(ws, "変更履歴", *META_SS)
    ws["B4"] = "変更履歴"
    ws["B4"].font = SECTION_FONT
    headers = ["No", "版", "日付", "更新者", "変更内容"]
    cw = [4, 6, 12, 8, 50]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    rev = [
        (1, "0.1", "2026-04-22", "佐藤", "初版作成 (受注登録モジュール)"),
        (2, "0.5", "2026-04-24", "佐藤", "金額計算 (税込換算) 追加、承認フロー定義"),
        (3, "0.7", "2026-04-25", "佐藤", "削除モジュール追加 (論理削除方針)"),
        (4, "0.8", "2026-04-26", "佐藤", "PT 項目記載、UI 1.2 反映"),
    ]
    for r, row in enumerate(rev, start=7):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c <= 3 else LEFT
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === 処理仕様 (受注登録 ORDS100) ===
    ws = wb.create_sheet("処理仕様_ORDS100")
    add_doc_header(ws, "処理仕様 (受注登録)", *META_SS)

    sub = [
        ("プロセスID", "ORDS100",  "プロセス名", "受注登録処理"),
        ("ジョブID",   "ORDS100_1", "ジョブ名",   "受注ヘッダ+明細INSERT"),
        ("対応UI",     "ORDC100",  "対応機能",   "受注登録画面"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(sub):
        r = 4 + i
        for col, val, font, fill in [
            (1, l1, META_FONT, META_FILL),
            (2, v1, MONO_FONT, None),
            (4, l2, META_FONT, META_FILL),
            (5, v2, BODY_FONT, None),
        ]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = font; cell.border = BORDER
            if fill: cell.fill = fill
            cell.alignment = CENTER

    headers = ["No", "処理概要", "詳細処理仕様"]
    cw = [6, 26, 60]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    blocks = [
        ("(1)", "メッセージ内容チェック",
         "リクエストされた電文の形式が正しいかチェックする\n"
         "※画面項目転送仕様(受注登録):登録 参照\n"
         "正しくない場合は MSG_003 を返却し終了"),
        ("(2)", "顧客マスタ存在チェック",
         "取得対象 : 顧客コード, 顧客名, 担当者メールアドレス\n"
         "参照テーブル: TBCTHB01 顧客マスタ\n"
         "取得条件 : CUSTOMER_CD = 引数.顧客コード\n"
         "         AND DELETE_FLG = '0'\n"
         "データが取得できない場合エラー設定:\n"
         "  メッセージID: MSG_009\n"
         "  「顧客マスタに存在しない顧客コードです」"),
        ("(3)", "明細毎の商品マスタ存在チェック + 入力検証",
         "for each 明細行:\n"
         "  取得対象 : 商品名, 単価, 税区分\n"
         "  参照テーブル: TBCTHB02 商品マスタ\n"
         "  取得条件: ITEM_CD = 明細.商品コード\n"
         "  if 数量 == null OR 数量 <= 0:\n"
         "    return MSG_004「数量は1以上で入力してください」\n"
         "  if 単価 == null:\n"
         "    return MSG_003「必須項目を入力してください」\n"
         "    ★UI で単価の必須/任意指定なし → 暫定で必須扱い"),
        ("(4)", "金額計算",
         "for each 明細行:\n"
         "  小計 = 数量 × 単価\n"
         "  if 税区分 == '8':\n"
         "    小計税込 = ROUND(小計 × 1.08, 0)\n"
         "  elif 税区分 == '10':\n"
         "    小計税込 = ROUND(小計 × 1.10, 0)\n"
         "  ★税区分 == 'その他' のケース未定義 (既定 10% 想定)\n"
         "合計金額 = SUM(小計税込)"),
        ("(5)", "承認フラグ判定",
         "if 合計金額 > 100000:\n"
         "  APPROVE_FLG = '0' (未承認)\n"
         "  STATUS = '20' (承認待ち)\n"
         "  メッセージID: MSG_006「承認待ちとなりました」\n"
         "else:\n"
         "  APPROVE_FLG = '1' (済)\n"
         "  STATUS = '30' (承認済)"),
        ("(6)", "受注ヘッダ INSERT",
         "更新対象 : ORDER_NO, CUSTOMER_CD, ORDER_DATE, TOTAL_AMOUNT, APPROVE_FLG, STATUS, DELETE_FLG\n"
         "更新テーブル: TBCTHB03 受注ヘッダ\n"
         "更新内容 :\n"
         "  ORDER_NO     = 'ORD' || YYYYMMDD || SEQ3 (採番関数 GetNextOrderNo)\n"
         "  CUSTOMER_CD  = 引数.顧客コード\n"
         "  ORDER_DATE   = システム日付\n"
         "  TOTAL_AMOUNT = (4)で計算した合計金額\n"
         "  APPROVE_FLG  = (5)で判定した値\n"
         "  STATUS       = (5)で判定した値\n"
         "  DELETE_FLG   = '0'\n"
         "INSERT 失敗時はトランザクション ROLLBACK + MSG_510 返却"),
        ("(7)", "受注明細 INSERT",
         "for each 明細行:\n"
         "  更新対象 : ORDER_NO, LINE_NO, ITEM_CD, QUANTITY, UNIT_PRICE, TAX_RATE, AMOUNT\n"
         "  更新テーブル: TBCTHB04 受注明細\n"
         "  LINE_NO は 1 から連番\n"
         "INSERT 失敗時はトランザクション ROLLBACK"),
        ("(8)", "メール通知キュー登録 (承認待ちのみ)",
         "if STATUS == '20' (承認待ち):\n"
         "  更新対象: 通知種別='APPROVAL_REQUEST', 受注番号, 担当者メール\n"
         "  更新テーブル: TBCTHB06 通知キュー\n"
         "  ORDB400 メール通知バッチが30秒間隔でポーリング → 送信"),
        ("(9)", "イベント履歴の作成",
         "イベント履歴登録用構造体を作成する\n"
         "  initialize_HistoryEvent で初期化\n"
         "  イベント区分 : 11 (受注登録)\n"
         "  対象キー    : 受注番号\n"
         "  実施ユーザ  : セッションユーザコード\n"
         "  実施日時    : システム日時\n"
         "イベント履歴登録 (TBCTHB99 へ INSERT)"),
        ("(10)", "画面側へレスポンスを返却する",
         "正常終了の場合:\n"
         "  STATUS='30' なら MSG_001「登録しました」\n"
         "  STATUS='20' なら MSG_006「承認待ちとなりました」\n"
         "エラーがある場合:\n"
         "  エラー発生箇所で設定したメッセージID を返却"),
    ]
    r = 9
    for no, summary, detail in blocks:
        ws.cell(row=r, column=1, value=no).font = BODY_BOLD
        ws.cell(row=r, column=1).alignment = CENTER; ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=summary).font = BODY_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP; ws.cell(row=r, column=2).border = BORDER
        cell = ws.cell(row=r, column=3, value=detail)
        cell.font = MONO_FONT
        cell.alignment = LEFT_TOP; cell.border = BORDER
        # 行の高さは内容の改行数に応じて
        nl = detail.count("\n")
        ws.row_dimensions[r].height = max(22, nl * 13 + 22)
        r += 1
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === 処理仕様 (削除 ORDS130) ===
    ws = wb.create_sheet("処理仕様_ORDS130")
    add_doc_header(ws, "処理仕様 (削除)", *META_SS)
    sub = [
        ("プロセスID", "ORDS130", "プロセス名", "受注削除処理"),
        ("対応UI",     "ORDC130", "対応機能",   "受注削除画面"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(sub):
        r = 4 + i
        for col, val, font, fill in [
            (1, l1, META_FONT, META_FILL),
            (2, v1, MONO_FONT, None),
            (4, l2, META_FONT, META_FILL),
            (5, v2, BODY_FONT, None),
        ]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = font; cell.border = BORDER
            if fill: cell.fill = fill
            cell.alignment = CENTER

    for i, h in enumerate(["No", "処理概要", "詳細処理仕様"], 1):
        c = ws.cell(row=8, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER

    del_blocks = [
        ("(1)", "受注ヘッダ存在チェック",
         "取得対象 : ORDER_NO, STATUS, DELETE_FLG\n"
         "参照テーブル: TBCTHB03 受注ヘッダ\n"
         "取得条件: ORDER_NO = 引数.受注番号\n"
         "         AND DELETE_FLG = '0'\n"
         "  ※for update を付与する\n"
         "存在しない場合: MSG_404「対象の受注が存在しません」"),
        ("(2)", "削除可能ステータスチェック",
         "if STATUS == '30' (承認済):\n"
         "  削除不可 → MSG_403「承認済の受注は削除できません」\n"
         "★UI 仕様で削除可能ステータスの明記なし (要確認)"),
        ("(3)", "受注ヘッダ 論理削除",
         "更新対象 : DELETE_FLG, STATUS, UPDATED_BY, UPDATED_AT\n"
         "更新テーブル: TBCTHB03 受注ヘッダ\n"
         "更新内容 :\n"
         "  DELETE_FLG = '1'\n"
         "  STATUS     = '90' (削除)\n"
         "  ※受注明細 (TBCTHB04) は物理削除しない (論理参照のため残置)\n"
         "  ★『適切に処理する』の具体: 在庫戻し処理は別途運用 (本処理では未対応)"),
        ("(4)", "イベント履歴 INSERT",
         "イベント区分 : 19 (受注削除)\n"
         "対象キー    : 受注番号\n"
         "実施ユーザ  : セッションユーザコード"),
        ("(5)", "戻り値",
         "正常時: MSG_002「削除しました」\n"
         "エラー時: 設定済みエラーメッセージ"),
    ]
    r = 9
    for no, summary, detail in del_blocks:
        ws.cell(row=r, column=1, value=no).font = BODY_BOLD
        ws.cell(row=r, column=1).alignment = CENTER; ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=summary).font = BODY_BOLD
        ws.cell(row=r, column=2).alignment = LEFT_TOP; ws.cell(row=r, column=2).border = BORDER
        cell = ws.cell(row=r, column=3, value=detail)
        cell.font = MONO_FONT
        cell.alignment = LEFT_TOP; cell.border = BORDER
        nl = detail.count("\n")
        ws.row_dimensions[r].height = max(22, nl * 13 + 22)
        r += 1
    for i, w in enumerate([6, 26, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === PT項目 ===
    ws = wb.create_sheet("PT項目")
    add_doc_header(ws, "PT項目 (チェック仕様)", *META_SS)
    ws["B4"] = "PT項目 (処理仕様 ORDS100/ORDS130 対応)"
    ws["B4"].font = SECTION_FONT
    headers = ["PT-ID", "対応プロセス", "対応No", "種別", "前提条件", "操作", "期待結果"]
    cw = [9, 14, 8, 8, 22, 20, 30]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    rows = [
        ("PT-001", "ORDS100", "(2)",  "正常", "顧客コード=ABC001 が存在", "受注登録実行",  "顧客名が画面に自動セット"),
        ("PT-002", "ORDS100", "(2)",  "異常", "顧客コード=XXX が不存在",  "受注登録実行",  "MSG_009 表示"),
        ("PT-003", "ORDS100", "(3)",  "正常", "数量=2, 単価=5000",       "金額計算",      "小計=10000, 税込=11000"),
        ("PT-004", "ORDS100", "(3)",  "異常", "数量=0",                   "受注登録",      "MSG_004 表示"),
        ("PT-005", "ORDS100", "(4)",  "境界", "税区分=8, 小計=12345",     "税込計算",      "ROUND(12345*1.08)=13333"),
        ("PT-006", "ORDS100", "(5)",  "境界", "合計=100000",              "受注登録",      "STATUS=30 承認済"),
        ("PT-007", "ORDS100", "(5)",  "境界", "合計=100001",              "受注登録",      "STATUS=20 承認待ち, MSG_006"),
        ("PT-008", "ORDS100", "(6)",  "正常", "正常データ",               "受注登録",      "TBCTHB03 INSERT 成功"),
        ("PT-009", "ORDS100", "(8)",  "正常", "STATUS=20",                "受注登録",      "TBCTHB06 通知キュー INSERT"),
        ("PT-010", "ORDS130", "(2)",  "異常", "STATUS=30 (承認済)",       "削除実行",      "MSG_403 表示, DELETE_FLG変化なし"),
        ("PT-011", "ORDS130", "(3)",  "正常", "STATUS=10 (登録)",         "削除実行",      "DELETE_FLG=1, STATUS=90 更新"),
    ]
    for r, row in enumerate(rows, start=7):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            cell.alignment = CENTER if c <= 4 else LEFT
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = OUT / "sample_SS_構造設計書.xlsx"
    wb.save(out)
    return out


# ============================================================================
# 観点ファイル 3種 (前回維持、追加列なし)
# ============================================================================
def _gen_aspect(filename, scope, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "観点"
    headers = ["観点名", "対象", "重要度", "工程", "レビュー指示文 (AIへのプロンプト)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = BORDER
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.alignment = LEFT_TOP; cell.border = BORDER
    cw = [22, 10, 8, 12, 60]
    for i, w in enumerate(cw, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    out = OUT / filename
    wb.save(out)
    return out


def gen_aspect_ui():
    return _gen_aspect("sample_観点_UI.xlsx", "UI", [
        ("必須項目欠落", "UI", "高", "基本設計",
         "機能概要書の画面定義/データ項目で、必須/任意の指定が欠落している項目を検出してください。"),
        ("曖昧表現", "UI", "中", "基本設計",
         "「適切に処理」「等」「など」のような曖昧な表現を検出し、後工程での解釈ブレ防止のため具体化を提案してください。"),
        ("用語不統一", "UI", "中", "基本設計",
         "同一概念に対して異なる用語が使われている箇所を検出してください (例: 受注 vs オーダー)。"),
        ("業務ロジック粒度不足", "UI", "中", "基本設計",
         "業務ルール記述で、条件分岐や境界値が明示されていない箇所を検出してください。"),
        ("メッセージ網羅性", "UI", "中", "基本設計",
         "正常系のみメッセージ定義され、異常系メッセージが不足している処理を検出してください。"),
        ("顧客向け表記", "UI", "低", "基本設計",
         "技術用語 (NULL, FK, バリデーション等) が顧客向け文書に紛れ込んでいないか確認してください。"),
    ])


def gen_aspect_ss():
    return _gen_aspect("sample_観点_SS.xlsx", "SS", [
        ("条件分岐網羅", "SS", "高", "詳細設計",
         "if/else/null/境界値で、想定すべき条件が漏れている箇所を検出してください。"),
        ("データ型整合", "SS", "高", "詳細設計",
         "UI で指定した型 (数値/文字/ブール) と SS の処理ロジックの型扱いが一致しているか確認してください。"),
        ("PT項目1:1対応", "SS", "高", "詳細設計",
         "SS の各処理ロジック行に対して PT項目が1対1で存在しているか、欠けている処理を検出してください。"),
        ("エラーハンドリング", "SS", "中", "詳細設計",
         "例外発生時の戻り値・ロールバック・ログ出力が記述されていない処理を検出してください。"),
        ("メッセージID整合", "SS", "中", "詳細設計",
         "SS で参照しているメッセージ ID が UI のメッセージ定義に存在するか確認してください。"),
        ("コメント/補足情報", "SS", "低", "詳細設計",
         "複雑な計算ロジックで、計算根拠や式の意図がコメント記載されていない箇所を検出してください。"),
    ])


def gen_aspect_cross():
    return _gen_aspect("sample_観点_整合性.xlsx", "UI×SS", [
        ("UI機能のSS反映漏れ", "UI×SS", "高", "詳細設計",
         "UI の機能一覧に存在するが、SS の処理仕様に対応する記述が無い機能を検出してください。"),
        ("SSにあるUI外処理", "UI×SS", "高", "詳細設計",
         "SS には記述されているが、UI に対応する機能/業務処理が無い処理を検出してください。"),
        ("仕様の食い違い", "UI×SS", "高", "詳細設計",
         "同一機能で UI と SS の処理内容や数値・条件が食い違っている箇所を検出してください (例: UIは税込/SSは税抜)。"),
        ("用語不統一(UI↔SS)", "UI×SS", "中", "詳細設計",
         "UI と SS で同一概念を異なる用語で表現している箇所を検出してください。"),
        ("メッセージID不整合", "UI×SS", "中", "詳細設計",
         "UI のメッセージ定義と SS の処理ロジックで参照しているメッセージID が一致しない箇所を検出してください。"),
        ("PT項目の網羅性", "UI×SS", "高", "詳細設計",
         "UI の業務ロジックや異常系に対して、SS の PT項目で網羅されていない条件を検出してください。"),
    ])


# ============================================================================
# 実行
# ============================================================================
if __name__ == "__main__":
    files = [
        gen_ui(),
        gen_ss(),
        gen_aspect_ui(),
        gen_aspect_ss(),
        gen_aspect_cross(),
    ]
    print("生成完了 (本格 SI 品質、受領フォーマット精査反映):")
    for f in files:
        size = f.stat().st_size
        print(f"  ✓ {f.name}  ({size:,} bytes)")
